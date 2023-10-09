import sys
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import( 
QApplication, QMainWindow, QPushButton, QLabel, QLineEdit)
from PyQt5.QtGui import QPixmap
from PyQt5.QtGui import QColor
from openpyxl import load_workbook
from functools import partial
# ~ import form9
import tokenf9
# ~ from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QWidget, QLabel
def ssinit():
		global wx0, wy0, wxl, wyl
		wx0, wy0, wxl, wyl = 200, 200, 1200, 800
		global L, W, nL, E, nE, ws, wb, FI, nB, B, nR, R
		nL = 24
		L = []
		nE = nL
		E = []
		FI = 2
		B = []
		nB = 4
		R = []
		nR = 6
	# ~ global wg
	# ~ wg = QRect(wx0, wy0, wxl, wyl)
# Subclass QMainWindow to customize your application's main window
class MainWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle("My App")
		self.setGeometry(wx0, wy0, wxl, wyl)

def sslabels():
	for i in range (nL):
		L.append (QLabel( parent = W))
		# ~ L[i].setText (str(i+10) )
		L[i].setGeometry(920- (i//12) * 500, 100 + (i%12)*50, 180, 40)
		L[i].setText(ws.cell(1, i+1).value)

def ssedits():
	for i in range (nE):
		E.append (QLineEdit( parent = W))
		# ~ E[i].setText (str(i+10) )
		E[i].setGeometry(800 - 500 * (i//12), 100 + (i% 12)*50, 120, 40)
		E[i].setText(str(ws.cell(2, i + 1).value))
		E[i].textChanged.connect(partial(textchanged, i))

	E[0].setEnabled(False)
	E[1].setEnabled(False)
	E[2].setEnabled(False)
	E[nE - 1].setEnabled(False)
	E[0].setStyleSheet("QLineEdit {color: black; }")
	E[1].setStyleSheet("QLineEdit {color: black; }")
	E[2].setStyleSheet("QLineEdit {color: black; }")
	E[nE - 1].setStyleSheet("QLineEdit {color: black; }")

def sseditsset():
	for i in range (nE):
		ss = str(ws.cell(FI, i+1).value)
		if ss == 'None':
			ss = ''
		E[i].setText(ss)
def editstocells():
	for i in range (nE):
		ws.cell(FI, i+1).value = E[i].text()
		print((E[i].text()))
	sseditsset()
def ssaction(i):
	global FI, wb
	# ~ print (B[i].text(), FI)
	if i == 0:
		FI += 1
	if i == 1:
		FI -= 1
	if FI <= 2:
		FI = 2
	if FI > ws.max_row:
		FI = ws.max_row
	# ~ print (FI)
	sseditsset()
	if i == 2:
		tokenf9.token9(ws, FI)
	if i == 3:
		editstocells()
#		pass
		wb.save(ename)
		
def textchanged (i):
	global FI, wb, ws
	ws.cell(FI, i + 1).value = str(E[i].text())
	#E[i].setStyleSheet("QLineEdit {color: red; }")
	print('44', FI, i)
	#wb.save(ename)

def ssbuttons():
	for i in range (nB):
		B.append (QPushButton( parent = W))
		B[i].setText (str(i+10) )
		B[i].setGeometry(100, 100 + i*50, 120, 40)
		B[i].clicked.connect (partial(ssaction, i))
	B[0].setText('next')
	B[1].setText('previous')
	B[2].setText('صدور توکن شارژ')
	B[3].setText('اصلاح پرونده')
def ssradios():
	for i in range (nL):
		R.append (QLabel( parent = W))
		L[i].setGeometry(920- (i//11) * 500, 100 + (i%11)*50, 180, 40)
		L[i].setText(ws.cell(1, i+1).value)

ssinit()

app = QApplication(sys.argv)
W = MainWindow()
ename = '20.xlsx'
wb = load_workbook(filename = ename)
ws = wb.active
print (ws.max_row)
print (ws.max_column)
print(ws.cell(1,1).value)
wdata = list(ws.values)
print (wdata)
for value in wdata:
	print (value)
# ~ form9.form9()
# ~ ss = QLabel( parent = W)
# ~ ss.setText ('salammmmmmmmdegjrjwjig' )
# ~ ss.setGeometry(wx0,wy0, wxl, wyl)
# ~ ss.setPixmap(QPixmap('dw.png'))

# ~ ss.setPixmap(pixmap_image)

sslabels()
ssedits()
sseditsset()
ssbuttons()
W.show()
app.exec_()
