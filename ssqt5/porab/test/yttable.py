from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import Qt

import sys
import openpyxl

class Main(QWidget):
	def __init__(self):
		super().__init__()
		self.setWindowTitle('Load From Excel')
		
		layout = QVBoxLayout()
		self.setLayout(layout)
		
		self.tablewidget = QTableWidget()
		layout.addWidget(self.tablewidget)
		
		self.load_data()
	def load_data(self):
		path = 	'/home/saeed/wp/ssqt5/porab/test/t1.xlsx'
		wb = openpyxl.load_workbook(filename =path)
		sheet = wb.active
		
		self.tablewidget.setRowCount (sheet.max_row * 30)
		self.tablewidget.setColumnCount (sheet.max_column * 10)
		values = list(sheet.values)
		i = 0
		for value in values:
			print (value)
			j = 0
			if i == 0:
				self.tablewidget.setHorizontalHeaderLabels(list (value))
			else:
				for v in value:
					w = QTableWidgetItem (str(v))
					w.setTextAlignment(Qt.AlignCenter)

					self.tablewidget.setItem(i-1, j, w )
					j += 1
			i += 1
if __name__ == '__main__':
	app = QApplication(sys.argv)
	app.setLayoutDirection(Qt.RightToLeft)           # <-------------------

	window = Main()
	window.showMaximized()
	app.exec_()
